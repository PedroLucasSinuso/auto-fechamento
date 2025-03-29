from openpyxl import load_workbook
import re
import streamlit as st
from config import PAYMENT_MAPPINGS, DATE_FORMAT

# Carrega as planilhas com e sem VBA
def load_workbooks(sheet):
    wb_data_only = load_workbook(sheet, data_only=True)
    wb = load_workbook(sheet, keep_vba=True)
    return wb_data_only, wb

# Desativa a proteção das planilhas para permitir edição
def disable_protection(ws_saldo_caixa, ws_rel_fechamento):
    ws_saldo_caixa.protection.disable()
    ws_rel_fechamento.protection.disable()

# Atualiza os valores de caixa e Credsystem na planilha
def update_cash_and_credsystem(ws_saldo_caixa, ws_saldo_caixa_data_only, collect, change_requested, cash_fund):
    ws_saldo_caixa["F10"] = cash_fund
    ws_saldo_caixa["F12"] = ws_saldo_caixa_data_only["F29"].value
    ws_saldo_caixa["F14"] = ws_saldo_caixa_data_only["F30"].value

    # Inicializa as células zeradas
    ws_saldo_caixa["F16"] = None
    ws_saldo_caixa["F18"] = None
    ws_saldo_caixa["F20"] = None
    ws_saldo_caixa["F22"] = None

    # Atualiza as células caso haja troco ou coleta
    if collect:
        ws_saldo_caixa["F16"] = ws_saldo_caixa_data_only["F29"].value
        ws_saldo_caixa["F18"] = ws_saldo_caixa_data_only["F30"].value

    if change_requested:
        ws_saldo_caixa["F20"] = change_requested
        ws_saldo_caixa["F22"] = change_requested

# Limpa células específicas nas planilhas
def clear_cells(ws_rel_fechamento, ws_saldo_caixa):
    for cell_range in ["B13:B20", "B24:B25", "G13:G23", "G25:G26"]:
        for row in ws_rel_fechamento[cell_range]:
            for cell in row:
                cell.value = None

# Atualiza as datas nas planilhas
def update_dates(ws_rel_fechamento, ws_saldo_caixa, day):
    ws_rel_fechamento["B7"] = day.strftime(DATE_FORMAT)
    ws_saldo_caixa["G7"] = day.strftime(DATE_FORMAT)

# Inicia um novo dia na planilha, configurando valores iniciais
def start_new_day(sheet, day, collect, change_requested, cash_fund):
    try:
        wb_data_only, wb = load_workbooks(sheet)
        ws_saldo_caixa_data_only = wb_data_only["Saldo em Caixa"]
        ws_saldo_caixa = wb["Saldo em Caixa"]
        ws_rel_fechamento = wb["Rel. Fechamento de Caixa"]

        disable_protection(ws_saldo_caixa, ws_rel_fechamento)
        update_cash_and_credsystem(ws_saldo_caixa, ws_saldo_caixa_data_only, collect, change_requested, cash_fund)
        clear_cells(ws_rel_fechamento, ws_saldo_caixa)
        update_dates(ws_rel_fechamento, ws_saldo_caixa, day)

        return wb
    except Exception as e:
        error_message = f"Erro ao iniciar um novo dia: {e}"
        print(error_message)
        st.warning(error_message)
        return None

# Insere os valores de vendas brutas por terminal na planilha
def insert_terminal_values(ws_rel_fechamento, report):
    for i in range(1, 7):
        try:
            terminal_key = f"{i:03}"
            cell = f"B{12 + i}"
            if terminal_key in report['Gross_Sales']:
                ws_rel_fechamento[cell] = report['Gross_Sales'][terminal_key]
        except Exception as e:
            error_message = f"Erro ao inserir valor do terminal {terminal_key}: {e}"
            print(error_message)
            st.warning(error_message)

# Insere os valores totais de acréscimos e descontos na planilha
def insert_total_values(ws_rel_fechamento, report):
    try:
        ws_rel_fechamento["B19"] = report['Gross_Add']
    except Exception as e:
        error_message = f"Erro ao inserir valor total de acréscimos: {e}"
        print(error_message)
        st.warning(error_message)

    try:
        ws_rel_fechamento["B25"] = report['Discounts']
    except Exception as e:
        error_message = f"Erro ao inserir valor total de descontos: {e}"
        print(error_message)
        st.warning(error_message)

# Mapeia outras informações do relatório para células específicas na planilha
def map_other_information(ws_rel_fechamento, report):
    mappings = {
        'Exchanged_Items': "B24",
        'Shipping': "G23",
        'Expenses': "G25",
        'Credsystem': "G26",
        'Omnichannel': "G19",
        'Total_Cash_Outflow': "G25",
        'Total_Cash_Inflow': "G13"
    }
    for key, cell in mappings.items():
        try:
            if key in report:
                ws_rel_fechamento[cell] = report[key]
        except Exception as e:
            error_message = f"Erro ao mapear {key}: {e}"
            print(error_message)
            st.warning(error_message)

# Insere os valores dos métodos de pagamento na planilha
def insert_payment_methods(ws_rel_fechamento, report):
    for pattern, cell in PAYMENT_MAPPINGS.items():
        try:
            if any(re.search(pattern, item) for item in report['Payment_Methods']):
                match_item = next(item for item in report['Payment_Methods'] if re.search(pattern, item))
                ws_rel_fechamento[cell] = report['Payment_Methods'][match_item]
        except Exception as e:
            error_message = f"Erro ao inserir método de pagamento {pattern}: {e}"
            print(error_message)
            st.warning(error_message)

# Compara os totais de vendas e pagamentos para verificar consistência
def compare_totals(ws_rel_fechamento):
    def get_value(cell):
        return cell.value if cell.value is not None else 0

    gross_total_terminals = sum(get_value(ws_rel_fechamento[f"B{row}"]) for row in range(13, 23))
    net_sale = round(gross_total_terminals - (get_value(ws_rel_fechamento["B24"]) + get_value(ws_rel_fechamento["B25"])), 2)

    gross_total_payments = sum(get_value(ws_rel_fechamento[f"G{row}"]) for row in range(13, 23))
    net_pay = round(gross_total_payments - get_value(ws_rel_fechamento["G23"]), 2)

    if net_pay == net_sale:
        st.success("Bateu! :sunglasses:")
    else:
        st.error("Divergente. Confere a tabela por favor :skull:")

# Edita a planilha com base no relatório gerado
def sheetEdit(sheet, report, day, collect=False, change_requested=0, cash_fund=0.0, observations=None):
    wb = start_new_day(sheet, day, collect, change_requested, cash_fund)
    if wb is None:
        error_message = "Erro ao iniciar um novo dia"
        print(error_message)
        st.warning(error_message)
        return None

    ws_rel_fechamento = wb["Rel. Fechamento de Caixa"]

    insert_terminal_values(ws_rel_fechamento, report)
    insert_total_values(ws_rel_fechamento, report)
    map_other_information(ws_rel_fechamento, report)
    insert_payment_methods(ws_rel_fechamento, report)
    compare_totals(ws_rel_fechamento)

    # Adiciona observações na célula B30, se fornecidas
    if observations:
        ws_rel_fechamento["B30"] = observations

    return wb