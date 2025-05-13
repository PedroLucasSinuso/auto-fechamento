from openpyxl import load_workbook
from core.entities.report import Report
from core.interfaces.sheet_writer import SheetWriter
from infrastructure.config.settings import PAYMENT_MAPPINGS, DATE_FORMAT
from datetime import datetime
from typing import Tuple, Any
import re

class ExcelWriter(SheetWriter):
    def __init__(self):
        self.wb = None
        self.wb_data_only = None
    
    def load_workbooks(self, sheet_path: str):
        self.wb_data_only = load_workbook(sheet_path, data_only=True)
        self.wb = load_workbook(sheet_path, keep_vba=True)
        return self.wb
    
    def disable_protection(self, *sheets):
        for sheet in sheets:
            sheet.protection.disable()
    
    def update_cash_and_credsystem(self, ws_saldo_caixa, ws_saldo_caixa_data_only, 
                                 collect: bool, change_requested: float, cash_fund: float):
        ws_saldo_caixa["F10"] = cash_fund
        ws_saldo_caixa["F12"] = ws_saldo_caixa_data_only["F29"].value
        ws_saldo_caixa["F14"] = ws_saldo_caixa_data_only["F30"].value

        ws_saldo_caixa["F16"] = None
        ws_saldo_caixa["F18"] = None
        ws_saldo_caixa["F20"] = None
        ws_saldo_caixa["F22"] = None

        if collect:
            ws_saldo_caixa["F16"] = ws_saldo_caixa_data_only["F29"].value
            ws_saldo_caixa["F18"] = ws_saldo_caixa_data_only["F30"].value

        if change_requested:
            ws_saldo_caixa["F20"] = change_requested
            ws_saldo_caixa["F22"] = change_requested
    
    def clear_cells(self, ws_rel_fechamento):
        for cell_range in ["B13:B20", "B24:B25", "G13:G23", "G25:G26"]:
            for row in ws_rel_fechamento[cell_range]:
                for cell in row:
                    cell.value = None
    
    def update_dates(self, ws_rel_fechamento, ws_saldo_caixa, day: datetime):
        ws_rel_fechamento["B7"] = day.strftime(DATE_FORMAT)
        ws_saldo_caixa["G7"] = day.strftime(DATE_FORMAT)
    
    def insert_terminal_values(self, ws_rel_fechamento, report: Report):
        for i in range(1, 7):
            terminal_key = f"{i:03}"
            cell = f"B{12 + i}"
            if terminal_key in report.gross_sales:
                ws_rel_fechamento[cell] = report.gross_sales[terminal_key]
    
    def insert_total_values(self, ws_rel_fechamento, report: Report):
        ws_rel_fechamento["B19"] = report.gross_add
        ws_rel_fechamento["B25"] = report.discounts
    
    def map_other_information(self, ws_rel_fechamento, report: Report):
        mappings = {
            'exchanged_items': "B24",
            'shipping': "G23",
            'expenses': "G25",
            'credsystem': "G26",
            'omnichannel': "G19",
            'total_cash_outflow': "G25",
            'total_cash_inflow': "G13"
        }
        
        for key, cell in mappings.items():
            if hasattr(report, key):
                ws_rel_fechamento[cell] = getattr(report, key)
    
    def insert_payment_methods(self, ws_rel_fechamento, report: Report):
        for pattern, cell in PAYMENT_MAPPINGS.items():
            for method in report.payment_methods:
                if re.search(pattern, method):
                    ws_rel_fechamento[cell] = report.payment_methods[method]
                    break
    
    def compare_totals(self, ws_rel_fechamento) -> bool:
        def get_value(cell):
            return cell.value if cell.value is not None else 0

        gross_total_terminals = sum(get_value(ws_rel_fechamento[f"B{row}"]) for row in range(13, 23))
        net_sale = round(gross_total_terminals - (get_value(ws_rel_fechamento["B24"]) + get_value(ws_rel_fechamento["B25"])), 2)

        gross_total_payments = sum(get_value(ws_rel_fechamento[f"G{row}"]) for row in range(13, 23))
        net_pay = round(gross_total_payments - get_value(ws_rel_fechamento["G23"]), 2)

        return net_pay == net_sale
    
    def write(self, sheet_path: str, report: Report, day: datetime, 
             collect: bool = False, change_requested: float = 0, 
             cash_fund: float = 0.0) -> Tuple[Any, bool]:
        try:
            self.load_workbooks(sheet_path)
            
            ws_saldo_caixa_data_only = self.wb_data_only["Saldo em Caixa"]
            ws_saldo_caixa = self.wb["Saldo em Caixa"]
            ws_rel_fechamento = self.wb["Rel. Fechamento de Caixa"]
            
            self.disable_protection(ws_saldo_caixa, ws_rel_fechamento)
            self.update_cash_and_credsystem(ws_saldo_caixa, ws_saldo_caixa_data_only, collect, change_requested, cash_fund)
            self.clear_cells(ws_rel_fechamento)
            self.update_dates(ws_rel_fechamento, ws_saldo_caixa, day)
            
            self.insert_terminal_values(ws_rel_fechamento, report)
            self.insert_total_values(ws_rel_fechamento, report)
            self.map_other_information(ws_rel_fechamento, report)
            self.insert_payment_methods(ws_rel_fechamento, report)
            
            return self.wb, self.compare_totals(ws_rel_fechamento)
            
        except Exception as e:
            raise Exception(f"Erro ao escrever na planilha: {e}")